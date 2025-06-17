"""
Output Generators for Functional Clinical Reporting

This module provides output generation functionality for different formats,
mirroring the SAS RRG approach to generating RTF, PDF, and other output formats.
"""

from typing import Dict, List, Optional, Any, Union
import pandas as pd
from pathlib import Path
from datetime import datetime
import warnings

from .config import FunctionalConfig
from .table_builder import TableResult


class OutputGeneratorFactory:
    """Factory for creating output generators."""
    
    @staticmethod
    def create(format_type: str, config: FunctionalConfig) -> 'BaseOutputGenerator':
        """
        Create output generator for specified format.
        
        Parameters
        ----------
        format_type : str
            Output format ('rtf', 'pdf', 'html', 'excel')
        config : FunctionalConfig
            Configuration for output generation
            
        Returns
        -------
        BaseOutputGenerator
            Output generator instance
        """
        format_type = format_type.lower()
        
        if format_type == 'rtf':
            return RTFGenerator(config)
        elif format_type == 'pdf':
            return PDFGenerator(config)
        elif format_type == 'html':
            return HTMLGenerator(config)
        elif format_type == 'excel':
            return ExcelGenerator(config)
        else:
            raise ValueError(f"Unsupported output format: {format_type}")


class BaseOutputGenerator:
    """Base class for output generators."""
    
    def __init__(self, config: FunctionalConfig):
        """
        Initialize output generator.
        
        Parameters
        ----------
        config : FunctionalConfig
            Configuration for output generation
        """
        self.config = config
        
    def generate(self, table_result: TableResult, output_dir: Path) -> Path:
        """
        Generate output file.
        
        Parameters
        ----------
        table_result : TableResult
            Table result to generate output for
        output_dir : Path
            Output directory
            
        Returns
        -------
        Path
            Path to generated output file
        """
        raise NotImplementedError("Subclasses must implement generate method")
    
    def _get_output_filename(self, table_result: TableResult, extension: str) -> str:
        """Generate output filename."""
        table_type = table_result.metadata.get('table_type', 'table')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f"{table_type}_{timestamp}.{extension}"


class RTFGenerator(BaseOutputGenerator):
    """
    RTF output generator (equivalent to SAS RRG RTF output).
    
    Generates Rich Text Format files suitable for regulatory submissions.
    """
    
    def generate(self, table_result: TableResult, output_dir: Path) -> Path:
        """Generate RTF output file."""
        
        output_file = output_dir / self._get_output_filename(table_result, 'rtf')
        
        # Generate RTF content
        rtf_content = self._generate_rtf_content(table_result)
        
        # Write to file
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(rtf_content)
        
        return output_file
    
    def _generate_rtf_content(self, table_result: TableResult) -> str:
        """Generate RTF content for table."""
        
        # RTF header
        rtf_lines = [
            r'{\rtf1\ansi\deff0',
            r'{\fonttbl{\f0 Times New Roman;}}',
            r'{\colortbl;\red0\green0\blue0;}',
            r'\paperw12240\paperh15840',  # Letter size
            r'\margl1440\margr1440\margt1440\margb1440',  # 1 inch margins
            r'\f0\fs20',  # Times New Roman, 10pt
            ''
        ]
        
        # Title
        if table_result.title:
            rtf_lines.extend([
                r'\qc\b\fs24',  # Center, bold, 12pt
                self._escape_rtf(table_result.title),
                r'\b0\fs20\ql\par',  # Reset formatting
                r'\par'
            ])
        
        # Subtitle
        if table_result.subtitle:
            rtf_lines.extend([
                r'\qc\fs20',  # Center, 10pt
                self._escape_rtf(table_result.subtitle),
                r'\ql\par',  # Reset alignment
                r'\par'
            ])
        
        # Table
        if not table_result.data.empty:
            rtf_lines.extend(self._generate_rtf_table(table_result.data))
        
        # Footnotes
        if table_result.footnotes:
            rtf_lines.extend([
                r'\par',
                r'\fs16',  # 8pt for footnotes
            ])
            
            for i, footnote in enumerate(table_result.footnotes, 1):
                rtf_lines.append(f"{i}. {self._escape_rtf(footnote)}\\par")
        
        # RTF footer
        rtf_lines.append('}')
        
        return '\n'.join(rtf_lines)
    
    def _generate_rtf_table(self, data: pd.DataFrame) -> List[str]:
        """Generate RTF table content."""
        
        if data.empty:
            return [r'\par No data available\par']
        
        rtf_lines = []
        
        # Calculate column widths (simplified)
        num_cols = len(data.columns)
        col_width = 12240 // num_cols  # Distribute page width evenly
        
        # Table header
        rtf_lines.append(r'\trowd\trgaph108')
        
        # Define cell borders and widths
        for i in range(num_cols):
            cell_right = (i + 1) * col_width
            rtf_lines.append(f'\\clbrdrt\\brdrs\\clbrdrl\\brdrs\\clbrdrb\\brdrs\\clbrdrr\\brdrs\\cellx{cell_right}')
        
        # Header row
        rtf_lines.append(r'\b')  # Bold
        for col in data.columns:
            rtf_lines.append(f'{self._escape_rtf(str(col))}\\cell')
        rtf_lines.extend([r'\b0\row'])  # End bold, end row
        
        # Data rows
        for _, row in data.iterrows():
            # Row definition
            rtf_lines.append(r'\trowd\trgaph108')
            for i in range(num_cols):
                cell_right = (i + 1) * col_width
                rtf_lines.append(f'\\clbrdrt\\brdrs\\clbrdrl\\brdrs\\clbrdrb\\brdrs\\clbrdrr\\brdrs\\cellx{cell_right}')
            
            # Cell content
            for value in row:
                rtf_lines.append(f'{self._escape_rtf(str(value))}\\cell')
            rtf_lines.append(r'\row')
        
        return rtf_lines
    
    def _escape_rtf(self, text: str) -> str:
        """Escape special characters for RTF."""
        if pd.isna(text):
            return ''
        
        text = str(text)
        # Escape RTF special characters
        text = text.replace('\\', '\\\\')
        text = text.replace('{', '\\{')
        text = text.replace('}', '\\}')
        
        return text


class PDFGenerator(BaseOutputGenerator):
    """
    PDF output generator.
    
    Generates PDF files using reportlab or similar library.
    """
    
    def generate(self, table_result: TableResult, output_dir: Path) -> Path:
        """Generate PDF output file."""
        
        output_file = output_dir / self._get_output_filename(table_result, 'pdf')
        
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            
            # Create PDF document
            doc = SimpleDocTemplate(
                str(output_file),
                pagesize=letter,
                rightMargin=inch,
                leftMargin=inch,
                topMargin=inch,
                bottomMargin=inch
            )
            
            # Build content
            story = []
            styles = getSampleStyleSheet()
            
            # Title
            if table_result.title:
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Title'],
                    fontSize=14,
                    spaceAfter=12,
                    alignment=1  # Center
                )
                story.append(Paragraph(table_result.title, title_style))
            
            # Subtitle
            if table_result.subtitle:
                subtitle_style = ParagraphStyle(
                    'CustomSubtitle',
                    parent=styles['Normal'],
                    fontSize=10,
                    spaceAfter=12,
                    alignment=1  # Center
                )
                story.append(Paragraph(table_result.subtitle, subtitle_style))
            
            # Table
            if not table_result.data.empty:
                # Prepare table data
                table_data = [list(table_result.data.columns)]
                for _, row in table_result.data.iterrows():
                    table_data.append([str(val) for val in row])
                
                # Create table
                pdf_table = Table(table_data)
                pdf_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                story.append(pdf_table)
            
            # Footnotes
            if table_result.footnotes:
                story.append(Spacer(1, 12))
                footnote_style = ParagraphStyle(
                    'Footnote',
                    parent=styles['Normal'],
                    fontSize=8,
                    spaceAfter=6
                )
                
                for i, footnote in enumerate(table_result.footnotes, 1):
                    story.append(Paragraph(f"{i}. {footnote}", footnote_style))
            
            # Build PDF
            doc.build(story)
            
        except ImportError:
            # Fallback if reportlab not available
            warnings.warn("reportlab not available, generating text file instead")
            output_file = output_file.with_suffix('.txt')
            self._generate_text_fallback(table_result, output_file)
        
        return output_file
    
    def _generate_text_fallback(self, table_result: TableResult, output_file: Path):
        """Generate text fallback when PDF libraries not available."""
        
        with open(output_file, 'w', encoding='utf-8') as f:
            if table_result.title:
                f.write(f"{table_result.title}\n")
                f.write("=" * len(table_result.title) + "\n\n")
            
            if table_result.subtitle:
                f.write(f"{table_result.subtitle}\n\n")
            
            if not table_result.data.empty:
                f.write(table_result.data.to_string(index=False))
                f.write("\n\n")
            
            if table_result.footnotes:
                f.write("Footnotes:\n")
                for i, footnote in enumerate(table_result.footnotes, 1):
                    f.write(f"{i}. {footnote}\n")


class HTMLGenerator(BaseOutputGenerator):
    """HTML output generator."""
    
    def generate(self, table_result: TableResult, output_dir: Path) -> Path:
        """Generate HTML output file."""
        
        output_file = output_dir / self._get_output_filename(table_result, 'html')
        
        html_content = self._generate_html_content(table_result)
        
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        return output_file
    
    def _generate_html_content(self, table_result: TableResult) -> str:
        """Generate HTML content for table."""
        
        html_lines = [
            '<!DOCTYPE html>',
            '<html>',
            '<head>',
            '<meta charset="utf-8">',
            '<title>Clinical Report Table</title>',
            '<style>',
            'body { font-family: "Times New Roman", serif; margin: 40px; }',
            'h1 { text-align: center; font-size: 14pt; margin-bottom: 10px; }',
            'h2 { text-align: center; font-size: 10pt; margin-bottom: 20px; }',
            'table { border-collapse: collapse; width: 100%; margin: 20px 0; }',
            'th, td { border: 1px solid black; padding: 8px; text-align: left; }',
            'th { background-color: #f0f0f0; font-weight: bold; }',
            '.footnotes { font-size: 8pt; margin-top: 20px; }',
            '</style>',
            '</head>',
            '<body>'
        ]
        
        # Title
        if table_result.title:
            html_lines.append(f'<h1>{self._escape_html(table_result.title)}</h1>')
        
        # Subtitle
        if table_result.subtitle:
            html_lines.append(f'<h2>{self._escape_html(table_result.subtitle)}</h2>')
        
        # Table
        if not table_result.data.empty:
            html_lines.append('<table>')
            
            # Header
            html_lines.append('<thead><tr>')
            for col in table_result.data.columns:
                html_lines.append(f'<th>{self._escape_html(str(col))}</th>')
            html_lines.append('</tr></thead>')
            
            # Body
            html_lines.append('<tbody>')
            for _, row in table_result.data.iterrows():
                html_lines.append('<tr>')
                for value in row:
                    html_lines.append(f'<td>{self._escape_html(str(value))}</td>')
                html_lines.append('</tr>')
            html_lines.append('</tbody>')
            
            html_lines.append('</table>')
        
        # Footnotes
        if table_result.footnotes:
            html_lines.append('<div class="footnotes">')
            for i, footnote in enumerate(table_result.footnotes, 1):
                html_lines.append(f'<p>{i}. {self._escape_html(footnote)}</p>')
            html_lines.append('</div>')
        
        html_lines.extend(['</body>', '</html>'])
        
        return '\n'.join(html_lines)
    
    def _escape_html(self, text: str) -> str:
        """Escape HTML special characters."""
        if pd.isna(text):
            return ''
        
        text = str(text)
        text = text.replace('&', '&amp;')
        text = text.replace('<', '&lt;')
        text = text.replace('>', '&gt;')
        text = text.replace('"', '&quot;')
        text = text.replace("'", '&#x27;')
        
        return text


class ExcelGenerator(BaseOutputGenerator):
    """Excel output generator."""
    
    def generate(self, table_result: TableResult, output_dir: Path) -> Path:
        """Generate Excel output file."""
        
        output_file = output_dir / self._get_output_filename(table_result, 'xlsx')
        
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Clinical Table"
            
            current_row = 1
            
            # Title
            if table_result.title:
                ws.cell(row=current_row, column=1, value=table_result.title)
                ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
                ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')
                current_row += 2
            
            # Subtitle
            if table_result.subtitle:
                ws.cell(row=current_row, column=1, value=table_result.subtitle)
                ws.cell(row=current_row, column=1).font = Font(size=10)
                ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')
                current_row += 2
            
            # Table
            if not table_result.data.empty:
                # Headers
                for col_idx, col_name in enumerate(table_result.data.columns, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=str(col_name))
                    cell.font = Font(bold=True)
                    cell.border = Border(
                        top=Side(style='thin'),
                        bottom=Side(style='thin'),
                        left=Side(style='thin'),
                        right=Side(style='thin')
                    )
                
                current_row += 1
                
                # Data
                for _, row in table_result.data.iterrows():
                    for col_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=current_row, column=col_idx, value=str(value))
                        cell.border = Border(
                            top=Side(style='thin'),
                            bottom=Side(style='thin'),
                            left=Side(style='thin'),
                            right=Side(style='thin')
                        )
                    current_row += 1
            
            # Footnotes
            if table_result.footnotes:
                current_row += 1
                for i, footnote in enumerate(table_result.footnotes, 1):
                    ws.cell(row=current_row, column=1, value=f"{i}. {footnote}")
                    ws.cell(row=current_row, column=1).font = Font(size=8)
                    current_row += 1
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(output_file)
            
        except ImportError:
            # Fallback to CSV if openpyxl not available
            warnings.warn("openpyxl not available, generating CSV instead")
            output_file = output_file.with_suffix('.csv')
            table_result.data.to_csv(output_file, index=False)
        
        return output_file 